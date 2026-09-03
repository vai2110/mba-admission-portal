#!/usr/bin/env python3
"""
Independent MBA college content production agent.

This file intentionally does not import or execute any existing repository agent,
AGENTS.md, .github/agents, or .github/skills instructions.
"""

import csv, json, os, re, subprocess, time
from pathlib import Path
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "college-content-master.csv"
MODEL = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
API_KEY = os.getenv("GEMINI_API_KEY", "")
BATCH_SIZE = int(os.getenv("COLLEGE_BATCH_SIZE", "10"))
PUBLISH_THRESHOLD = 70
TIMEOUT = 20
session = requests.Session()
session.headers.update({"User-Agent": "MBA-College-Content-Agent/1.0"})

def slugify(s): return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")

def fetch(url):
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        if r.ok and "text/html" in r.headers.get("content-type", ""): return r.text, r.url
    except requests.RequestException: pass
    return "", url

def official_crawl(start_url, limit=24):
    domain = urlparse(start_url).netloc.lower().replace("www.", "")
    queue, seen, pages = [start_url], set(), []
    keys = ("mba","pgp","pgdm","management","admission","placement","career","fee","fees",
            "programme","program","curriculum","eligibility","selection","brochure","prospectus",
            "annual","report","school","department")
    while queue and len(pages) < limit:
        url = queue.pop(0)
        if url in seen or urlparse(url).netloc.lower().replace("www.", "") != domain: continue
        seen.add(url)
        html, final_url = fetch(url)
        if not html: continue
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(" ", strip=True)
        if text:
            pages.append({"url": final_url, "title": soup.title.get_text(" ", strip=True) if soup.title else "", "text": text[:18000]})
        for a in soup.find_all("a", href=True):
            href = urljoin(final_url, a["href"]).split("#")[0]
            if urlparse(href).netloc.lower().replace("www.", "") != domain: continue
            hay = (a.get_text(" ", strip=True) + " " + href).lower()
            if any(k in hay for k in keys) and href not in seen and href not in queue: queue.append(href)
    return pages

def gemini_json(prompt, max_tokens=30000):
    if not API_KEY: raise RuntimeError("GEMINI_API_KEY is missing")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent"
    payload = {"contents":[{"role":"user","parts":[{"text":prompt}]}],
               "generationConfig":{"temperature":0.2,"maxOutputTokens":max_tokens,"responseMimeType":"application/json"}}
    r = session.post(url, params={"key":API_KEY}, json=payload, timeout=180)
    r.raise_for_status()
    return json.loads(r.json()["candidates"][0]["content"]["parts"][0]["text"])

def research_block(college, official_url):
    pages = official_crawl(official_url)
    chunks = [f"SOURCE URL: {p['url']}\nTITLE: {p['title']}\nCONTENT:\n{p['text']}" for p in pages]
    return "\n\n--- OFFICIAL SOURCE ---\n\n".join(chunks), pages

def generation_prompt(college, rank, official_url, research):
    return f"""
You are the independent MBA admissions content generator for an Indian education portal.
COLLEGE: {college}
NIRF 2025 RANK IN MASTER LIST: {rank}
OFFICIAL START URL: {official_url}

FACT RULE: Use ONLY the supplied official-source material for factual claims. Never invent or infer
fees, dates, seats, salaries, cutoffs, eligibility, rankings, recruiters, programme names or statistics.
If unavailable, state that the official source reviewed does not publish/confirm it. Reddit/Quora are
permitted only for cutoff information when an official cutoff is absent, and only if evidence is supplied.

DESIGN RULE: The approved reference pages define architecture, content depth, hierarchy and responsive
presentation. Use the shared college-page.css implementation. Do not copy wording or facts from benchmark
colleges. Use concise hero, quick facts, desktop On This Page sidebar, mobile On This Page control,
cards/tables/answer boxes where useful, FAQs and official sources. Answer first, student-centric, simple English.

SCOPE: Create one overview page, one placement page, and up to three high-intent MBA/PGDM/Executive MBA
or closely related management programme pages supported by official sources. Do not create unrelated UG,
PhD or certificate pages.

OUTPUT JSON:
{{"pages":[{{"filename":"slug.html","type":"overview|placement|programme","title":"...","html":"<complete HTML document>","source_urls":["official URL"]}}]}}

HTML: complete HTML5; reference /college-page.css; one H1; logical H2/H3; title/meta description/canonical/
OG metadata/JSON-LD; FAQs; official source links; internal links only to known package/existing files;
no markdown, placeholders, lorem ipsum or unsupported numbers.

OFFICIAL SOURCE MATERIAL:
{research[:180000]}
"""

def audit_html(html, source_urls, all_repo_files):
    soup = BeautifulSoup(html, "html.parser"); text = soup.get_text(" ", strip=True).lower()
    score, reasons, critical = 0, [], []
    if soup.find("h1"): score += 5
    else: critical.append("missing H1")
    h2 = len(soup.find_all("h2")); score += min(15, h2*3)
    if h2 < 3: reasons.append("thin heading structure")
    if soup.find("meta", attrs={"name":"description"}): score += 5
    else: critical.append("missing meta description")
    if soup.find("link", rel="canonical"): score += 5
    else: critical.append("missing canonical")
    if soup.find("script", attrs={"type":"application/ld+json"}): score += 5
    else: reasons.append("missing JSON-LD")
    if soup.find("table"): score += 5
    if len(soup.find_all(class_=re.compile(r"faq", re.I))) >= 3 or text.count("faq") >= 3: score += 5
    if soup.find(string=re.compile("official source", re.I)): score += 5
    if len(source_urls) >= 2: score += 10
    elif source_urls: score += 6
    else: critical.append("no official sources")
    internal = [a.get("href","") for a in soup.find_all("a", href=True) if not a.get("href","").startswith(("http://","https://","#","mailto:"))]
    broken = [x for x in internal if x and Path(ROOT / x).name not in all_repo_files and not x.startswith("/")]
    if not broken: score += 10
    else: critical.append("broken internal link(s)")
    if "lorem" in text or "placeholder" in text: critical.append("placeholder content")
    if any(x in text for x in ("₹0","rs. 0","to be updated","insert here")): critical.append("placeholder/zero value")
    if len(text) >= 1800: score += 10
    elif len(text) >= 1000: score += 6
    else: reasons.append("low content depth")
    if soup.find("meta", attrs={"property":"og:title"}): score += 3
    if soup.find("meta", attrs={"property":"og:description"}): score += 2
    if "college-page.css" in str(html): score += 5
    return min(score,100), critical, reasons

def read_master():
    with MASTER.open(encoding="utf-8-sig", newline="") as f: return list(csv.DictReader(f))

def write_master(rows):
    with MASTER.open("w", encoding="utf-8-sig", newline="") as f:
        w=csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)

def update_xlsx(rows):
    path=ROOT/"college-production-tracker.xlsx"
    if not path.exists(): return
    wb=load_workbook(path); ws=wb.active; headers=[c.value for c in ws[1]]
    for h in ("Quality Score","Research Status","QA Status","Deployment Status","Live Verification"):
        if h not in headers: ws.cell(1,ws.max_column+1,h); headers.append(h)
    idx={h:i+1 for i,h in enumerate(headers)}
    for r in rows:
        for rr in range(2,ws.max_row+1):
            if str(ws.cell(rr,1).value or "").strip()==r["college_name"].strip():
                for h in ("Quality Score","Research Status","QA Status","Deployment Status","Live Verification"): ws.cell(rr,idx[h],r[h.lower().replace(" ","_")])
    wb.save(path)

def git(*args): subprocess.run(["git",*args],cwd=ROOT,check=True)

def main():
    rows=read_master(); existing={p.name for p in ROOT.glob("*.html")}
    eligible=[r for r in rows if any(not str(r.get(k," ")).strip().lower()=="done" for k in ("overview_status","placement_status","popular_course_status"))][:BATCH_SIZE]
    if not eligible: print("No eligible colleges."); return
    for r in eligible:
        college=r["college_name"].strip(); r["research_status"]="Researching"; write_master(rows); update_xlsx(rows)
        print(f"=== {r['rank']}: {college} ===")
        research, source_pages=research_block(college,r["official_website"]); r["research_status"]="Research Complete"; write_master(rows); update_xlsx(rows)
        data=gemini_json(generation_prompt(college,r["rank"],r["official_website"],research)); pages=data.get("pages",[])
        safe=[]
        for p in pages:
            fn=Path(p.get("filename","")).name
            if fn and fn not in existing and not (ROOT/fn).exists(): safe.append((fn,p))
        if not safe:
            r["qa_status"]="Failed - no new pages"; r["quality_score"]="0"; write_master(rows); update_xlsx(rows); continue
        package={fn for fn,_ in safe}
        for fn,p in safe:
            soup=BeautifulSoup(p["html"],"html.parser")
            if package:
                for target,tp in safe:
                    if target==fn: continue
                    a=soup.new_tag("a",href=target,attrs={"class":"official-link"}); a.string=tp.get("title",target)
                    if soup.body: soup.body.append(a)
            p["html"]=str(soup)
        scores=[]; critical_all=[]; notes=[]
        for fn,p in safe:
            sc,cr,no=audit_html(p["html"],p.get("source_urls",[]),existing|package); scores.append(sc)
            critical_all += [f"{fn}: {x}" for x in cr]; notes += [f"{fn}: {x}" for x in no]; (ROOT/fn).write_text(p["html"],encoding="utf-8")
        package_score=round(sum(scores)/len(scores)); r["quality_score"]=str(package_score); r["qa_status"]="Passed" if package_score>PUBLISH_THRESHOLD and not critical_all else "Failed"
        r["deployment_status"]="Not Started"; r["live_verification"]="Not Started"; report={"college":college,"rank":r["rank"],"score":package_score,"page_scores":scores,"critical_failures":critical_all,"notes":notes,"source_count":len(source_pages),"generated_pages":[x[0] for x in safe]}
        (ROOT/f"quality-audit-{slugify(college)}.json").write_text(json.dumps(report,indent=2),encoding="utf-8"); write_master(rows); update_xlsx(rows)
        if package_score>PUBLISH_THRESHOLD and not critical_all:
            r["deployment_status"]="Eligible - Auto Publish"; write_master(rows); update_xlsx(rows)
            git("add",*[fn for fn,_ in safe],"data/college-content-master.csv","college-production-tracker.xlsx",f"quality-audit-{slugify(college)}.json")
            git("commit","-m",f"Publish independently generated pages for {college}"); git("push","origin","main")
            r["deployment_status"]="Pushed - Awaiting Pages"; write_master(rows); update_xlsx(rows)
            urls=[]
            for fn,_ in safe:
                url=f"https://vai2110.github.io/mba-admission-portal/{fn}"; ok=False
                for _ in range(6):
                    try:
                        if session.get(url,timeout=20).status_code==200: ok=True; break
                    except requests.RequestException: pass
                    time.sleep(10)
                urls.append(ok)
            r["live_verification"]="Verified" if all(urls) else "Failed"; write_master(rows); update_xlsx(rows)
            git("add","data/college-content-master.csv","college-production-tracker.xlsx"); git("commit","-m",f"Update quality and live status for {college}"); git("push","origin","main")
            existing.update(package)
        else:
            r["deployment_status"]="Blocked - QA <= 70 or critical failure"; write_master(rows); update_xlsx(rows)
            git("add","data/college-content-master.csv","college-production-tracker.xlsx",f"quality-audit-{slugify(college)}.json",*[fn for fn,_ in safe]); git("commit","-m",f"Hold {college} pages for QA revision"); git("push","origin","main"); existing.update(package)

if __name__=="__main__": main()
