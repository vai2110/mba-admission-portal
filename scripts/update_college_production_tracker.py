import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parents[1]
QUEUE=ROOT/'data/college-queue.csv'; STATE=ROOT/'data/college-production-state.json'
OUT=ROOT/'college-production-tracker.xlsx'; CSV_OUT=ROOT/'college-production-tracker.csv'
PAGE_COLS=['Overview Page','Placement Page','Course Page 1','Course Page 2','Course Page 3']; DONE='Created + Audited'; PENDING='Pending'

# Keep the production queue's official/full names unchanged, but show the
# names students are more likely to recognise in the tracker/Google Sheet.
ALIASES={
    'Management Development Institute':'MDI Gurgaon',
    'XLRI - Xavier School of Management':'XLRI',
    'Symbiosis Institute of Business Management':'SIBM Pune',
    'Indian Institute of Foreign Trade':'IIFT',
    'S. P. Jain Institute of Management and Research':'SPJIMR',
    'Amrita Vishwa Vidyapeetham':'Amrita Vishwa Vidyapeetham',
    'Jamia Millia Islamia':'JMI',
    'Chandigarh University':'CU',
    'MICA':'MICA',
    'UPES':'UPES',
    'Great Lakes Institute of Management':'Great Lakes Institute of Management',
    'T. A. Pai Management Institute Manipal':'TAPMI Manipal',
    'IMI Delhi':'IMI Delhi',
    'IMI Kolkata':'IMI Kolkata',
    'Goa Institute of Management':'GIM',
    'Lovely Professional University':'LPU',
    'XIM University':'XIM University',
    'Thapar Institute of Engineering and Technology (Deemed-to-be-university)':'Thapar University',
    'Amity University':'Amity University',
    'Graphic Era University':'Graphic Era',
    'Nirma University':'Nirma University',
    'Institute of Rural Management Anand':'IRMA',
    'Loyola Institute of Business Administration':'LIBA',
    'S.R.M. Institute of Science and Technology':'SRMIST',
    'Christ University':'Christ University',
    'Fore School of Management':'FORE School of Management',
    'Banaras Hindu University':'BHU',
    'Birla Institute of Management Technology':'BIMTECH',
    'Malaviya National Institute of Technology':'MNIT Jaipur',
    'Saveetha Institute of Medical and Technical Sciences':'SIMATS',
    'K.J.Somaiya Institute of Management':'K J Somaiya Institute of Management',
    'Siksha `O` Anusandhan':'SOA University',
    'Kalinga Institute of Industrial Technology':'KIIT',
    'Aligarh Muslim University':'AMU',
    'Alliance University':'Alliance University',
    'Prin. L.N. Welingkar Institute of Management Development and Research (PGDM)':'Welingkar (WeSchool)',
    'Guru Gobind Singh Indraprastha University':'GGSIPU',
    'BML Munjal University':'BMU',
    'Chitkara University':'Chitkara University',
    'Babasheb Bhimrao Ambedkar University':'BBAU',
    'Thiagarajar School of Management':'TSM',
    'Manipal University Jaipur':'MUJ',
    'Cochin University of Science and Technology':'CUSAT',
    'Madan Mohan Malaviya University of Technology':'MMMUT',
    'PSG College of Technology':'PSG Tech',
    'New Delhi Institute of Management':'NDIM',
    'Jamia Hamdard':'Jamia Hamdard',
    'Anna University':'Anna University',
    'Pandit Deendayal Energy University':'PDEU',
    'Jagan Institute of Management Studies':'JIMS',
    'Rajagiri Business School':'RBS',
    'Panjab University':'PU',
    'Atal Bihari Vajpayee Indian Institute of Information Technology and Management':'ABV-IIITM Gwalior',
    'National Institute of Agricultural Extension Management':'MANAGE',
    'Bharathidasan Institute of Management':'BIM Trichy',
    'Birla Institute of Technology':'BIT Mesra',
    'University of Lucknow':'Lucknow University',
}

def display_name(r):
    name=r['college_name'].strip(); url=r.get('official_url','').lower(); rank=int(r['rank'])
    if name.startswith('Indian Institute of Management'):
        if rank==25 or 'iimnagpur' in url: return 'IIM Nagpur'
        return name.replace('Indian Institute of Management, Mumbai','IIM Mumbai').replace('Indian Institute of Management, Amritsar','IIM Amritsar').replace('Indian Institute of Management Jammu (IIMJ)','IIM Jammu').replace('Indian Institute of Management ','IIM ')
    if name.startswith('Indian Institute of Technology'):
        if 'Indian School of Mines' in name: return 'IIT (ISM) Dhanbad'
        return name.replace('Indian Institute of Technology ','IIT ')
    if name.startswith('National Institute of Technology'):
        return name.replace('National Institute of Technology Tiruchirappalli','NIT Trichy').replace('National Institute of Technology ','NIT ')
    if name=='Institute of Management Technology':
        return 'IMT Nagpur' if 'imtnagpur' in url else 'IMT Ghaziabad'
    if 'Institute of Management Technology, Nagpur' in name: return 'IMT Nagpur'
    if 'Great Lakes Institute of Management' in name:
        return 'Great Lakes Gurgaon' if 'greatlakesgurgaon' in url else 'Great Lakes Chennai'
    if 'Jaipuria Institute of Management, Lucknow' in name: return 'Jaipuria Lucknow'
    if name=='Jaipuria Institute of Management' and 'noida' in url: return 'Jaipuria Noida'
    if name=='Jaipuria Institute of Management': return 'Jaipuria Institute of Management'
    if name=='SVKM`s Narsee Monjee Institute of Management Studies': return 'NMIMS'
    if name=='ICFAI Foundation for Higher Education, Hyderabad': return 'ICFAI Business School Hyderabad'
    if name=='Koneru Lakshmaiah Education Foundation University (K L College of Engineering)': return 'KL University'
    if name=='Jain university,Bangalore': return 'JAIN University'
    return ALIASES.get(name,name)

def loadq():
    with QUEUE.open(encoding='utf-8-sig',newline='') as f: return list(csv.DictReader(f))
def loads():
    try: return json.loads(STATE.read_text(encoding='utf-8'))
    except Exception: return {}
def rows():
    q=loadq(); s=loads(); legacy={int(x) for x in s.get('legacy_completed_ranks',list(range(1,22)))}; out=[]
    for r in q:
        rank=int(r['rank']); rec=s.get('colleges',{}).get(str(rank),{}); vals=[rec.get(c,DONE if rank in legacy else PENDING) for c in PAGE_COLS]
        out.append([display_name(r),*vals])
    return out

def build():
    data=rows(); headers=['College Name',*PAGE_COLS]
    with CSV_OUT.open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.writer(f); w.writerow(headers); w.writerows(data)
    wb=Workbook(); ws=wb.active; ws.title='College Queue'; ws.append(headers)
    for r in data: ws.append(r)
    out=wb.create_sheet('Completed Outside 100'); out.append(['College Name','Status','Pages']); out.append(['SIBM Bengaluru',DONE,'Overview; MBA; MBA Business Analytics; MBA Executive; Placements'])
    legend=wb.create_sheet('Status Legend'); legend.append(['Status','Meaning'])
    for a,b in [(PENDING,'Not started'),('Creating','Currently being created'),('Created — Audit Pending','Created but audit is not yet passed'),(DONE,'Created and audit passed'),('Needs Review','Blocked or failed validation')]: legend.append([a,b])
    navy='17336F'; green='15803D'; amber='B45309'; blue='2563EB'; red='B91C1C'; gray='64748B'; border='DCE3EB'; white='FFFFFF'; thin=Side(style='thin',color=border)
    for sh in wb.worksheets:
        sh.sheet_view.showGridLines=False
        for c in sh[1]: c.font=Font(bold=True,color=white); c.fill=PatternFill('solid',fgColor=navy); c.alignment=Alignment(wrap_text=True,vertical='center'); c.border=Border(bottom=thin)
        sh.row_dimensions[1].height=28
        for row in sh.iter_rows(min_row=2):
            for c in row: c.alignment=Alignment(wrap_text=True,vertical='top'); c.border=Border(bottom=thin)
    for i,w in enumerate([42,24,24,24,24,24],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='B2'; ws.auto_filter.ref=f'A1:F{ws.max_row}'
    from openpyxl.formatting.rule import FormulaRule
    rng=f'B2:F{ws.max_row}'
    for formula,fill,font in [('B2="Created + Audited"','DCFCE7',green),('B2="Pending"','F8FAFC',gray),('B2="Creating"','FEF3C7',amber),('B2="Created — Audit Pending"','DBEAFE',blue),('B2="Needs Review"','FEE2E2',red)]: ws.conditional_formatting.add(rng,FormulaRule(formula=[formula],fill=PatternFill('solid',fgColor=fill),font=Font(color=font,bold=True)))
    wb.save(OUT)

if __name__=='__main__': build()
