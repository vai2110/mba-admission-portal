import json
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "college-page-audit-dashboard.xlsx"
EXCLUDED = {"index.html", "404.html", "content-audit.html", "college-page-audit-dashboard.html"}

NAVY="17336F"; BLUE="2563EB"; GREEN="15803D"; AMBER="B45309"; GRAY="64748B"; BORDER="DCE3EB"; WHITE="FFFFFF"
THIN=Side(style="thin",color=BORDER)

def run(cmd):
    return subprocess.check_output(cmd,cwd=ROOT,text=True,stderr=subprocess.DEVNULL).strip()

def read_json(name):
    p=ROOT/name
    try: return json.loads(p.read_text(encoding="utf-8"))
    except Exception: return {}

def report_item(report, filename):
    for x in report.get("results",[]):
        if x.get("file")==filename: return x
    return {}

def college_name(path):
    stem=Path(path).stem
    parts=stem.split("-")
    if parts[:2]==["iim","rohtak"]: return "IIM Rohtak"
    if parts[:2]==["iim","udaipur"]: return "IIM Udaipur"
    if parts[:2]==["iim","mumbai"]: return "IIM Mumbai"
    return " ".join(x.upper() if len(x)<=3 else x.title() for x in parts[:3])

def page_type(path):
    s=Path(path).stem
    if "placement" in s: return "Placement"
    if s.count("-")<=2: return "Overview"
    return "Course-specific"

def today_pages():
    out=run(["git","log","--since=00:00","--name-only","--pretty=format:","--","*.html"])
    return sorted({x.strip() for x in out.splitlines() if x.strip().endswith(".html") and x.strip() not in EXCLUDED and not x.startswith("sibm-pune") and not x.startswith("iim-ahmedabad")})

def make():
    content=read_json("content-audit-report.json"); seo=read_json("seo-geo-aeo-report.json")
    pages=today_pages(); rows=[]
    for p in pages:
        c=report_item(content,p); s=report_item(seo,p); audited=bool(c or s)
        cscore=c.get("overall_score"); sscore=s.get("overall_score")
        vals=[v for v in (cscore,sscore) if isinstance(v,(int,float))]
        score=round(sum(vals)/len(vals)) if vals else None
        status="AUDITED" if audited else "IN QUEUE"
        modified="MODIFIED / VERIFIED" if audited and (c.get("auto_applied") or s.get("auto_applied")) else ("AUDITED – NO CHANGE" if audited else "Awaiting audit")
        word_count=c.get("metrics",{}).get("word_count",0)
        notes=[]
        if c.get("issues"): notes.append(f"Content audit: {len(c['issues'])} issue group(s)")
        if s.get("issues"): notes.append(f"SEO/AEO/GEO audit: {len(s['issues'])} issue group(s)")
        if not notes: notes.append("Awaiting benchmark-level verification.")
        rows.append([college_name(p),p,page_type(p),status,modified,"Review", "Pass" if word_count>=1200 else "Review", "Pass" if s else "Review", "Pass" if s else "Review", "Pass" if s else "Review", "Review","Review","Review"," ".join(notes),run(["git","log","-1","--format=%H","--",p]),score])

    wb=Workbook(); dash=wb.active; dash.title="Live Dashboard"
    dash["A1"]="College Page Audit & Modification Dashboard"; dash["A1"].font=Font(size=18,bold=True,color=WHITE); dash["A1"].fill=PatternFill("solid",fgColor=NAVY); dash.merge_cells("A1:F1")
    dash["A2"]="Live repository status · IIM Ahmedabad = content benchmark · SIBM Pune = architecture/design benchmark"; dash["A2"].font=Font(italic=True,color=GRAY); dash.merge_cells("A2:F2")
    stats=[("Pages today",len(rows)),("Audited",sum(r[3]=="AUDITED" for r in rows)),("In queue",sum(r[3]=="IN QUEUE" for r in rows)),("Modified / verified",sum("MODIFIED" in r[4] for r in rows))]
    for i,(label,val) in enumerate(stats):
        col=1+i*2; dash.cell(4,col,label).font=Font(bold=True,color=GRAY); dash.cell(5,col,val).font=Font(size=18,bold=True,color=BLUE); dash.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1); dash.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
    dash["A7"]="Last updated (UTC)"; dash["B7"]=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    dash["A8"]="End goal"; dash["B8"]="Every non-benchmark page reaches IIM Ahmedabad content depth + SIBM Pune architecture/design standard."; dash.merge_cells("B8:F8")

    aq=wb.create_sheet("Audit Queue"); headers=["College","Page","Page Type","Audit Status","Modification Status","Architecture","Content Depth","SEO","AEO","GEO","Official Sources","Internal Links","Placement/Data","Notes","Latest Commit","Audit Score"]; aq.append(headers)
    for r in rows: aq.append(r)
    if len(aq.rows)>1:
        t=Table(displayName="LiveAuditQueue",ref=aq.dimensions); t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); aq.add_table(t)
    aq.freeze_panes="A2"

    bm=wb.create_sheet("Benchmarks"); bm.append(["Benchmark","Area","Locked Standard"])
    for r in [["IIM Ahmedabad","Content depth","Detailed programme-specific admission logic, exact eligibility/selection criteria, official-source discipline, student-intent coverage and actual data."],["SIBM Pune","Architecture & design","Compact text-first hero, clean navigation, section hierarchy, cards/tables, responsive layout and no invented hero components."],["Both","SEO/AEO/GEO","Search-intent H1/H2s, long-tail coverage, direct answers, metadata/canonical, structured data, local/entity signals and internal linking."],["Both","Data integrity","Use official institute numbers; if unavailable, explicitly say unavailable rather than fabricate estimates."]]: bm.append(r)

    cl=wb.create_sheet("Audit Checklist"); cl.append(["Category","Check","Required Outcome"])
    for r in [["Architecture","Hero","SIBM Pune-style compact text-first hero; no invented logo/placement/stat cards."],["Architecture","Navigation","Consistent On-this-page navigation, hierarchy and internal buttons."],["Content","Admission","Detailed eligibility → application → exam/test → shortlist → PI/selection → final offer."],["Content","Fees","Fee bifurcation + actual course-period cost; additional student-incurred costs separated."],["Content","Cutoff","Official cutoff where published; otherwise clearly labelled expected/indicative without false precision."],["Content","Placements","Latest highlights + maximum 3 comparable years of actual placement statistics."],["Content","Specificity","No generic filler; programme/college-specific facts and student decision support."],["SEO","Metadata","Unique title, meta description, canonical, OG/Twitter and clean H1."],["AEO","Answer blocks","Direct answers to high-intent queries and useful non-generic FAQs."],["GEO","Entity/location","Institute, city/state and official identity clearly represented."],["Technical","Section separation","<hr/> between consecutive H2 sections where required."],["Technical","Responsive","Hero, cards, tables and navigation usable on mobile and desktop."],["Integrity","Sources","Factual claims grounded in official institute sources; unavailable data not invented."]]: cl.append(r)

    lg=wb.create_sheet("Change Log"); lg.append(["Timestamp UTC","College","Page","Audit Status","Modification Status","Commit"])
    for r in rows: lg.append([datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),r[0],r[1],r[3],r[4],r[14]])

    for sh in wb.worksheets:
        sh.sheet_view.showGridLines=False
        for cell in sh[1]: cell.font=Font(bold=True,color=WHITE); cell.fill=PatternFill("solid",fgColor=NAVY); cell.alignment=Alignment(vertical="center",wrap_text=True); cell.border=Border(bottom=THIN)
        for row in sh.iter_rows(min_row=2):
            for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True); cell.border=Border(bottom=THIN)
        sh.row_dimensions[1].height=28
    widths={"Live Dashboard":[24,32,20,20,20,20],"Audit Queue":[18,34,18,16,24,16,16,10,10,10,18,16,20,62,44,14],"Benchmarks":[18,25,100],"Audit Checklist":[18,30,100],"Change Log":[23,18,34,16,25,44]}
    for name,vals in widths.items():
        for i,w in enumerate(vals,1): wb[name].column_dimensions[get_column_letter(i)].width=w
    for row in aq.iter_rows(min_row=2):
        for cell in row:
            if cell.value in ("AUDITED","MODIFIED / VERIFIED","Pass"): cell.font=Font(color=GREEN,bold=True)
            elif cell.value in ("IN QUEUE","Awaiting audit","Review"): cell.font=Font(color=AMBER,bold=True)
    wb.save(OUT)

if __name__=="__main__": make()
