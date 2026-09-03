#!/usr/bin/env python3
"""Independent MBA college content production agent v5."""
import csv,json,os,re,subprocess,time
from pathlib import Path
from urllib.parse import urljoin,urlparse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1];MASTER=ROOT/"data"/"college-content-master.csv";PRIORITY=ROOT/"data"/"college-content-priority.csv";OVERRIDES=ROOT/"data"/"college-content-overrides.csv";TRACKER=ROOT/"college-production-tracker.csv"
MODEL=os.getenv("GEMINI_MODEL","gemini-3.1-flash-lite");API_KEY=os.getenv("GEMINI_API_KEY","");BATCH_SIZE=int(os.getenv("COLLEGE_BATCH_SIZE","10"));PUBLISH_THRESHOLD=70;TIMEOUT=20
session=requests.Session();session.headers.update({"User-Agent":"MBA-College-Content-Agent/1.0"})
def norm(s):return re.sub(r"[^a-z0-9]","",s.lower())
def slugify(s):return re.sub(r"[^a-z0-9]+","-",s.lower()).strip("-")
def fetch(url):
    try:
        r=session.get(url,timeout=TIMEOUT,allow_redirects=True)
        if r.ok and "text/html" in r.headers.get("content-type",""):return r.text,r.url
    except requests.RequestException:pass
    return "",url
def discover_official_url(college,supplied_url):
    if supplied_url:
        html,final=fetch(supplied_url)
        if html:return final
    try:
        r=requests.get("https://www.google.com/search",params={"q":f'"{college}" official website'},timeout=TIMEOUT,headers={"User-Agent":"Mozilla/5.0"});soup=BeautifulSoup(r.text,"html.parser")
        bad=("wikipedia.org","collegedunia.com","shiksha.com","careers360.com","collegedekho.com","getmyuni.com","facebook.com","linkedin.com")
        for a in soup.find_all("a",href=True):
            href=a["href"]
            if href.startswith("/url?q="):href=href.split("/url?q=",1)[1].split("&",1)[0]
            if href.startswith("http") and not any(x in href.lower() for x in bad):
                h,final=fetch(href)
                if h:return final
    except requests.RequestException:pass
    return ""
def official_crawl(start_url,limit=28):
    domain=urlparse(start_url).netloc.lower().replace("www.","");queue=[start_url];seen=set();pages=[];keys=("mba","pgp","pgdm","management","admission","placement","career","fee","fees","programme","program","curriculum","eligibility","selection","brochure","prospectus","annual","report","school","department","course")
    while queue and len(pages)<limit:
        url=queue.pop(0)
        if url in seen or urlparse(url).netloc.lower().replace("www.","")!=domain:continue
        seen.add(url);html,final=fetch(url)
        if not html:continue
        soup=BeautifulSoup(html,"html.parser");text=soup.get_text(" ",strip=True)
        if text:pages.append({"url":final,"title":soup.title.get_text(" ",strip=True) if soup.title else "","text":text[:18000]})
        for a in soup.find_all("a",href=True):
            href=urljoin(final,a["href"]).split("#")[0]
            if urlparse(href).netloc.lower().replace("www.","")!=domain:continue
            hay=(a.get_text(" ",strip=True)+" "+href).lower()
            if any(k in hay for k in keys) and href not in seen and href not in queue:queue.append(href)
    return pages
def research_block(url):
    pages=official_crawl(url);return "\n\n--- OFFICIAL SOURCE ---\n\n".join(f"SOURCE URL: {p['url']}\nTITLE: {p['title']}\nCONTENT:\n{p['text']}" for p in pages),pages
def gemini_json(prompt,max_tokens=30000):
    if not API_KEY:raise RuntimeError("GEMINI_API_KEY is missing")
    url=f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent";payload={"contents":[{"role":"user","parts":[{"text":prompt}]}],"generationConfig":{"temperature":0.15,"maxOutputTokens":max_tokens,"responseMimeType":"application/json"}}
    last=None
    for attempt in range(4):
        try:
            r=session.post(url,params={"key":API_KEY},json=payload,timeout=180)
            if r.status_code in (429,500,502,503,504):last=r;time.sleep(5*(attempt+1));continue
            r.raise_for_status();return json.loads(r.json()["candidates"][0]["content"]["parts"][0]["text"])
        except requests.RequestException as e:
            last=e
            if attempt<3:time.sleep(5*(attempt+1))
            else:raise
    raise RuntimeError(f"Gemini unavailable after retries: {last}")
def read_master():
    with MASTER.open(encoding="utf-8-sig",newline="") as f:return list(csv.DictReader(f))
def priority_rows(rows):
    if not PRIORITY.exists():return rows
    by_rank={str(r.get("rank")):r for r in rows};out=[]
    with PRIORITY.open(encoding="utf-8-sig",newline="") as f:
        for p in csv.DictReader(f):
            r=by_rank.get(str(p.get("rank")))
            if r:r["college_name"]=p["college_name"];r["official_website"]=p["official_website"];out.append(r)
    return out+[r for r in rows if r not in out]
def read_overrides():
    if not OVERRIDES.exists():return {}
    with OVERRIDES.open(encoding="utf-8-sig",newline="") as f:return {str(r.get("rank")):r for r in csv.DictReader(f)}
def apply_overrides(rows):
    ov=read_overrides()
    for r in rows:
        x=ov.get(str(r.get("rank")))
        if x:
            for k in ("overview_status","placement_status","popular_course_status","quality_score","research_status","qa_status","deployment_status","live_verification"):
                if x.get(k)!="":r[k]=x[k]
    return ov
def read_tracker():
    if not TRACKER.exists():return []
    with TRACKER.open(encoding="utf-8-sig",newline="") as f:return list(csv.DictReader(f))
def tracker_for(college,tracker):
    def keys(s):
        raw=s.lower();v={norm(raw)}
        for a,b in [("indian institute of management","iim"),("indian institute of technology","iit"),("national institute of technology","nit"),("institute of management technology","imt"),("management development institute","mdi"),("symbiosis institute of business management","sibm"),("t. a. pai management institute","tapmi"),("great lakes institute of management","great lakes"),("chandigarh university","cu")]:
            if a in raw:v.add(norm(raw.replace(a,b)))
        return v
    tk=keys(college)
    for r in tracker:
        if tk & keys(r.get("College Name","")):return r
    return None
def missing_types(college,tracker,forced=None):
    if forced:return forced
    r=tracker_for(college,tracker)
    if not r:return {"overview","placement","programme"}
    def created(v):return "created" in (v or "").lower() or "audited" in (v or "").lower()
    m=set()
    if not created(r.get("Overview Page")):m.add("overview")
    if not created(r.get("Placement Page")):m.add("placement")
    courses=[r.get(k,"") for k in ("Course Page 1","Course Page 2","Course Page 3")]
    if courses and not any(created(v) for v in courses) and not all((v or "").strip().lower()=="not applicable" for v in courses):m.add("programme")
    return m
def prompt(college,rank,url,research,types):
    return f'''Independent MBA admissions page generation. College: {college}. NIRF 2025 rank: {rank}. Official domain: {url}. Generate ONLY: {", ".join(sorted(types))}. Use ONLY supplied official-source material for facts. Never invent fees, dates, seats, salaries, cutoffs, eligibility, rankings, recruiters, programme names or statistics. If unavailable, say so. Reddit/Quora may be used only for cutoff information when official cutoff is absent and evidence is supplied. Use the approved IIM Ahmedabad/SIBM Pune reference architecture and shared college-page.css; do not copy their wording/facts. Student-centric, simple English, answer-first. Return JSON ONLY with pages [{filename,type,title,html,source_urls}]. Complete HTML5, one H1, logical H2/H3, title/meta description/canonical/OG/JSON-LD, FAQs, official sources, internal links only to known existing/package files. No placeholders, markdown or unsupported numbers.\n\nOFFICIAL SOURCE MATERIAL:\n{research[:180000]}'''
def revision_prompt(college,rank,url,research,types,failures,previous):
    return f'''Revise these MBA admissions pages for {college}. Use ONLY supplied official sources. Fix: {failures}. Preserve verified facts, remove unsupported claims, improve metadata, structure, FAQs, sources and responsive architecture. Generate ONLY {", ".join(sorted(types))}. Return JSON with complete HTML5 pages. PREVIOUS: {json.dumps(previous)[:120000]}\nSOURCES:\n{research[:160000]}'''
def audit(html,source_urls,official_url,files):
    soup=BeautifulSoup(html,"html.parser");text=soup.get_text(" ",strip=True).lower();score=0;critical=[];notes=[]
    if soup.find("h1"):score+=5
    else:critical.append("missing H1")
    h2=len(soup.find_all("h2"));score+=min(15,h2*3)
    if h2<3:notes.append("thin heading structure")
    if soup.find("meta",attrs={"name":"description"}):score+=5
    else:critical.append("missing meta description")
    if soup.find("link",rel="canonical"):score+=5
    else:critical.append("missing canonical")
    if soup.find("script",attrs={"type":"application/ld+json"}):score+=5
    else:notes.append("missing JSON-LD")
    if soup.find("table"):score+=5
    if len(soup.find_all(class_=re.compile(r"faq",re.I)))>=3 or text.count("faq")>=3:score+=5
    if soup.find(string=re.compile("official source",re.I)):score+=5
    if len(source_urls)>=2:score+=10
    elif source_urls:score+=6
    else:critical.append("no official sources")
    domain=urlparse(official_url).netloc.lower().replace("www.","")
    if any(urlparse(u).netloc.lower().replace("www.","")!=domain for u in source_urls):critical.append("non-official source URL")
    internal=[a.get("href","") for a in soup.find_all("a",href=True) if not a.get("href","").startswith(("http://","https://","#","mailto:"))]
    broken=[x for x in internal if x and Path(ROOT/x).name not in files and not x.startswith("/")]
    if not broken:score+=10
    else:critical.append("broken internal link(s)")
    if "lorem" in text or "placeholder" in text:critical.append("placeholder content")
    if any(x in text for x in ("₹0","rs. 0","to be updated","insert here")):critical.append("placeholder/zero value")
    if len(text)>=1800:score+=10
    elif len(text)>=1000:score+=6
    else:notes.append("low content depth")
    if soup.find("meta",attrs={"property":"og:title"}):score+=3
    if soup.find("meta",attrs={"property":"og:description"}):score+=2
    if "college-page.css" in str(html):score+=5
    return min(score,100),critical,notes
def write_master(rows):
    with MASTER.open("w",encoding="utf-8-sig",newline="") as f:w=csv.DictWriter(f,fieldnames=list(rows[0].keys()));w.writeheader();w.writerows(rows)
def update_xlsx(rows):
    path=ROOT/"data"/"college-content-master.xlsx"
    if path.exists():wb=load_workbook(path);ws=wb.active;ws.delete_rows(1,ws.max_row)
    else:
        from openpyxl import Workbook
        wb=Workbook();ws=wb.active;ws.title="Management"
    headers=list(rows[0].keys())
    for c,h in enumerate(headers,1):ws.cell(1,c,h)
    for rr,r in enumerate(rows,2):
        for c,h in enumerate(headers,1):ws.cell(rr,c,r.get(h,""))
    ws.freeze_panes="A2";wb.save(path)
def git(*args):subprocess.run(["git",*args],cwd=ROOT,check=True)
def main():
    rows=read_master();overrides=apply_overrides(rows);tracker=read_tracker();existing={p.name for p in ROOT.glob("*.html")};generated=0
    for r in priority_rows(rows):
        if generated>=BATCH_SIZE:break
        college=r["college_name"].strip();ov=overrides.get(str(r.get("rank")),{});forced={x.strip() for x in (ov.get("force_missing","") or "").split(",") if x.strip()};types=missing_types(college,tracker,forced)
        if not types:
            r["qa_status"]="Skipped - existing pages protected";r["deployment_status"]="Already Exists";r["live_verification"]="Checked";write_master(rows);update_xlsx(rows);continue
        r["research_status"]="Researching";write_master(rows);update_xlsx(rows);print(f"=== {r['rank']}: {college} | {sorted(types)} ===")
        url=discover_official_url(college,r["official_website"])
        if not url:
            r["research_status"]="Blocked - official site not discovered";r["qa_status"]="Failed";r["quality_score"]="0";write_master(rows);update_xlsx(rows);continue
        r["official_website"]=url;research,source_pages=research_block(url);r["research_status"]="Research Complete";write_master(rows);update_xlsx(rows)
        data=gemini_json(prompt(college,r["rank"],url,research,types));pages=data.get("pages",[]);safe=[]
        for p in pages:
            if p.get("type") not in types:continue
            fn=Path(p.get("filename","")).name
            if fn and fn not in existing and not (ROOT/fn).exists():safe.append((fn,p))
        if not safe:
            r["qa_status"]="Failed - generator returned no missing page";r["quality_score"]="0";write_master(rows);update_xlsx(rows);continue
        package={fn for fn,_ in safe};scores=[];critical_all=[];notes=[]
        for fn,p in safe:
            sc,cr,no=audit(p["html"],p.get("source_urls",[]),url,existing|package);scores.append(sc);critical_all += [f"{fn}: {x}" for x in cr];notes += [f"{fn}: {x}" for x in no];(ROOT/fn).write_text(p["html"],encoding="utf-8")
        package_score=round(sum(scores)/len(scores))
        if package_score<=PUBLISH_THRESHOLD or critical_all:
            revision=gemini_json(revision_prompt(college,r["rank"],url,research,types,critical_all+notes,pages));rev=[p for p in revision.get("pages",[]) if p.get("type") in types and p.get("html")]
            if rev:
                safe=[(Path(p.get("filename","")).name,p) for p in rev if Path(p.get("filename","")).name];package={fn for fn,_ in safe};scores=[];critical_all=[];notes=[]
                for fn,p in safe:
                    sc,cr,no=audit(p["html"],p.get("source_urls",[]),url,existing|package);scores.append(sc);critical_all += [f"{fn}: {x}" for x in cr];notes += [f"{fn}: {x}" for x in no];(ROOT/fn).write_text(p["html"],encoding="utf-8")
                package_score=round(sum(scores)/len(scores))
        r["quality_score"]=str(package_score);r["qa_status"]="Passed" if package_score>PUBLISH_THRESHOLD and not critical_all else "Failed"
        for k,t in (("overview_status","overview"),("placement_status","placement"),("popular_course_status","programme")):
            if any(p.get("type")==t for _,p in safe):r[k]="Done"
        report={"college":college,"rank":r["rank"],"score":package_score,"page_scores":scores,"critical_failures":critical_all,"notes":notes,"source_count":len(source_pages),"generated_pages":[x[0] for x in safe]};(ROOT/f"quality-audit-{slugify(college)}.json").write_text(json.dumps(report,indent=2),encoding="utf-8");r["deployment_status"]="Not Started";r["live_verification"]="Not Started";write_master(rows);update_xlsx(rows)
        git("config","user.email","41898282+github-actions[bot]@users.noreply.github.com");git("config","user.name","github-actions[bot]")
        if package_score>PUBLISH_THRESHOLD and not critical_all:
            r["deployment_status"]="Eligible - Auto Publish";write_master(rows);update_xlsx(rows);git("add",*[fn for fn,_ in safe],"data/college-content-master.csv","data/college-content-master.xlsx",f"quality-audit-{slugify(college)}.json");git("commit","-m",f"Publish independently generated pages for {college}");git("push","origin","main");r["deployment_status"]="Pushed - Awaiting Pages";write_master(rows);update_xlsx(rows);urls=[]
            for fn,_ in safe:
                url2=f"https://vai2110.github.io/mba-admission-portal/{fn}";ok=False
                for _ in range(8):
                    try:
                        if session.get(url2,timeout=20).status_code==200:ok=True;break
                    except requests.RequestException:pass
                    time.sleep(10)
                urls.append(ok)
            r["live_verification"]="Verified" if all(urls) else "Failed";write_master(rows);update_xlsx(rows);git("add","data/college-content-master.csv","data/college-content-master.xlsx");git("commit","-m",f"Update quality and live status for {college}");git("push","origin","main")
        else:
            r["deployment_status"]="Blocked - QA <= 70 or critical failure"
            for fn,_ in safe:
                try:(ROOT/fn).unlink()
                except FileNotFoundError:pass
            write_master(rows);update_xlsx(rows);git("add","data/college-content-master.csv","data/college-content-master.xlsx",f"quality-audit-{slugify(college)}.json");git("commit","-m",f"Hold {college} pages for QA revision");git("push","origin","main")
        existing.update(package);generated+=1
if __name__=="__main__":main()
