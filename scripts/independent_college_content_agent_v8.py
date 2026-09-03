#!/usr/bin/env python3
"""Standalone v8 MBA college page generator.

This agent deliberately does NOT read, import, execute, or follow AGENTS.md or
any other repository agent configuration. Reference pages are used only as
visual/content benchmarks supplied in this agent's own contract.

Core design decision:
Gemini generates STRUCTURED CONTENT, not final HTML. This script owns the
HTML architecture and renders every page through one deterministic template.
That prevents the model from changing layout, hero composition, card hierarchy,
sidebar placement, spacing, or responsive structure from college to college.
"""
import csv, json, os, re, subprocess, time
from html import escape
from pathlib import Path
from urllib.parse import urlencode, urljoin, urlparse
from urllib.request import Request, urlopen

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "college-content-master.csv"
OVERRIDES = ROOT / "data" / "college-content-overrides.csv"
MODEL = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
API_KEY = os.getenv("GEMINI_API_KEY", "")
SHEET_URL = os.getenv("GOOGLE_SHEET_WEBAPP_URL", "").strip()
SHEET_SECRET = os.getenv("GOOGLE_SHEETS_API_SECRET", "").strip()
BATCH_SIZE = int(os.getenv("COLLEGE_BATCH_SIZE", "10"))
TEST_RANK = os.getenv("COLLEGE_TEST_RANK", "").strip()
COMPLETED_UPTO_RANK = 26
PUBLISH_THRESHOLD = 70
MAX_REVISIONS = 2
TIMEOUT = 25
session = requests.Session()
session.headers.update({"User-Agent": "MBA-College-Content-Agent/3.0"})


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s or "").lower()).strip("-")


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def fetch(url):
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        if r.ok and "text/html" in r.headers.get("content-type", ""):
            return r.text, r.url
    except requests.RequestException:
        pass
    return "", url


def discover_official(college, supplied):
    if supplied:
        html, final = fetch(supplied)
        if html:
            return final
    return ""


def official_crawl(start, limit=35):
    domain = urlparse(start).netloc.lower().replace("www.", "")
    q, seen, pages = [start], set(), []
    keys = ("mba","pgp","pgdm","management","admission","placement","career","fee","fees","programme","program","curriculum","eligibility","selection","brochure","prospectus","annual","report","course","school")
    while q and len(pages) < limit:
        url = q.pop(0).split("#")[0]
        if url in seen or urlparse(url).netloc.lower().replace("www.", "") != domain:
            continue
        seen.add(url)
        html, final = fetch(url)
        if not html:
            continue
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(" ", strip=True)
        if text:
            pages.append({"url": final, "title": soup.title.get_text(" ", strip=True) if soup.title else "", "text": text[:20000]})
        for a in soup.find_all("a", href=True):
            href = urljoin(final, a["href"]).split("#")[0]
            if urlparse(href).netloc.lower().replace("www.", "") != domain:
                continue
            hay = (a.get_text(" ", strip=True) + " " + href).lower()
            if any(k in hay for k in keys) and href not in seen and href not in q:
                q.append(href)
    return pages


def research(start):
    pages = official_crawl(start)
    block = "\n\n--- OFFICIAL SOURCE ---\n\n".join(f"SOURCE URL: {p['url']}\nTITLE: {p['title']}\nCONTENT:\n{p['text']}" for p in pages)
    return block, pages


def gemini(prompt, tokens=30000):
    if not API_KEY:
        raise RuntimeError("GEMINI_API_KEY is missing")
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent"
    payload = {"contents":[{"role":"user","parts":[{"text":prompt}]}],"generationConfig":{"temperature":0.12,"maxOutputTokens":tokens,"responseMimeType":"application/json"}}
    for attempt in range(4):
        try:
            r = session.post(endpoint, params={"key":API_KEY}, json=payload, timeout=180)
            if r.status_code in (429,500,502,503,504):
                time.sleep(5*(attempt+1)); continue
            r.raise_for_status()
            return json.loads(r.json()["candidates"][0]["content"]["parts"][0]["text"])
        except requests.RequestException:
            if attempt == 3: raise
            time.sleep(5*(attempt+1))
    raise RuntimeError("Gemini request failed")


def read_master():
    with MASTER.open(encoding="utf-8-sig", newline="") as f: return list(csv.DictReader(f))


def write_master(rows):
    with MASTER.open("w", encoding="utf-8-sig", newline="") as f:
        w=csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)


def update_xlsx(rows):
    p=ROOT/"data"/"college-content-master.xlsx"
    if p.exists(): wb=load_workbook(p); ws=wb.active; ws.delete_rows(1, ws.max_row)
    else:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title="Management"
    for c,h in enumerate(rows[0].keys(),1): ws.cell(1,c,h)
    for rr,r in enumerate(rows,2):
        for c,h in enumerate(rows[0].keys(),1): ws.cell(rr,c,r.get(h,""))
    ws.freeze_panes="A2"; wb.save(p)


def read_overrides():
    if not OVERRIDES.exists(): return {}
    with OVERRIDES.open(encoding="utf-8-sig", newline="") as f: return {str(r.get("rank")):r for r in csv.DictReader(f)}


def sheet_get(action, **params):
    if not SHEET_URL: raise RuntimeError("GOOGLE_SHEET_WEBAPP_URL is missing")
    u=SHEET_URL+("&" if "?" in SHEET_URL else "?")+urlencode({"action":action,**params})
    req=Request(u,headers={"User-Agent":"MBA-College-Content-Agent/3.0"})
    with urlopen(req,timeout=60) as r: data=json.loads(r.read().decode())
    if not data.get("success"): raise RuntimeError(data.get("error","Google Sheet read failed"))
    return data


def sheet_post(action, **fields):
    if not SHEET_URL or not SHEET_SECRET: raise RuntimeError("Google Sheet credentials are missing")
    body=json.dumps({"secret":SHEET_SECRET,"action":action,**fields}).encode()
    req=Request(SHEET_URL,data=body,headers={"Content-Type":"application/json","User-Agent":"MBA-College-Content-Agent/3.0"},method="POST")
    with urlopen(req,timeout=60) as r: data=json.loads(r.read().decode())
    if not data.get("success"): raise RuntimeError(data.get("error","Google Sheet update failed"))
    return data


def eligible_from_sheet():
    data=sheet_get("nextBatch",batchSize=max(BATCH_SIZE,10))
    out=[]
    for x in data.get("colleges",[]):
        try: rank=int(str(x.get("rank",x.get("Rank",""))).strip())
        except: continue
        if rank<=COMPLETED_UPTO_RANK: continue
        out.append({"rank":str(rank),"college_name":str(x.get("collegeName",x.get("College Name",x.get("college_name","")))).strip(),"official_website":str(x.get("officialWebsite",x.get("official_website",x.get("Official Website Links","")))).strip(),"overview_status":str(x.get("overviewStatus",x.get("Overview Page",""))).strip(),"placement_status":str(x.get("placementStatus",x.get("Placement Page",""))).strip(),"popular_course_status":str(x.get("popularCourseStatus",x.get("Popular Course Pages",""))).strip()})
    if TEST_RANK:
        out=[x for x in out if x["rank"]==TEST_RANK]
    return out[:BATCH_SIZE]


def done(v): return str(v or "").strip().lower() in {"done","complete","completed","already exists","verified","published","live"}

def missing_types(meta):
    if meta.get("overview_status") and not done(meta.get("overview_status")): pass
    m=[]
    if not done(meta.get("overview_status")): m.append("overview")
    if not done(meta.get("placement_status")): m.append("placement")
    if not done(meta.get("popular_course_status")): m.append("programme")
    return m


def existing_files(): return {p.name for p in ROOT.glob("*.html")}


def internal_targets(existing, package_types, college):
    # Prefer known pages from the same college; then current package filenames.
    cs=slug(college); hits=sorted(x for x in existing if x.startswith(cs+"-"))
    return hits


def content_prompt(college, rank, url, research, types, existing):
    ordered=[x for x in ("overview","placement","programme") if x in types]
    return f"""You are the CONTENT RESEARCHER for an MBA admissions website. You are NOT the HTML designer. Return structured JSON content only.
COLLEGE: {college}
NIRF 2025 RANK: {rank}
OFFICIAL DOMAIN: {url}
PAGE TYPES REQUIRED, IN THIS ORDER: {ordered}

REFERENCE QUALITY CONTRACT:
The target experience follows two approved benchmarks: IIM Ahmedabad for information depth, answer-first student usefulness, tables and decision support; SIBM Pune for compact text-first hero, navigation, section rhythm, cards, tables, CTA placement and responsive behaviour. Do not copy their facts or wording. The HTML architecture will be rendered deterministically by the agent; you must provide rich, college-specific information for it.

CONTENT STANDARD:
- Write for a student deciding whether/how to apply, not an institutional brochure.
- Simple English, concise but information-rich.
- Every important student question must be answered when official evidence exists.
- Never fabricate or infer fees, dates, seats, salaries, recruiters, cutoffs, eligibility, rankings, programme names or statistics.
- If an important fact is not published in the supplied official material, return "Not published by the official source" rather than guessing.
- Distinguish qualifying cutoff from competitive/expected target. If official cutoff is absent, say so. Student-reported cutoff evidence is allowed only when supplied separately; do not invent it.
- Separate programme fee, deposits and other costs.
- Use latest official placement data only; never attribute institution-wide data to a programme unless the source does so.
- FAQs must be real student questions and college-specific.
- Include exact official source URLs for every material section.

REQUIRED PAGE ARCHITECTURE CONTENT:
OVERVIEW: overview; important admission dates; programmes/courses; eligibility; detailed admission process; entrance exam; selection criteria; cutoff; fees; fee breakup; total course-period cost; placements; popular programmes; student decision factors; FAQs.
PROGRAMME: quick answer; programme overview; important dates; eligibility; admission process; entrance exam; selection criteria; cutoff/qualifying threshold; fees; fee breakup; total cost; curriculum; programme structure; programme-specific placement/career information; who should apply/decision factors; FAQs.
PLACEMENT: latest placement highlights; placement snapshot; up to 3 comparable years; average CTC; median CTC; highest CTC; recruiters/offers where official; function/sector data where official; placement trends; programme-specific interpretation; FAQs.

Return JSON exactly as {{"pages":[...]}}. Each page object must contain type, title, hero_subtitle, location_line, facts (exactly 4 items with label/value), quick_answer, sections (ordered list), faqs (at least 5), cta_label, cta_url, source_urls. Each section has id,title,answer,paragraphs,bullets,table with columns and rows, callout. Keep empty arrays when genuinely not applicable. Do not put HTML in the response.

KNOWN EXISTING HTML FILES (only these may be linked):
{chr(10).join('- '+x for x in sorted(existing)[:300])}

OFFICIAL SOURCE MATERIAL:
{research[:220000]}"""


def normalize_content(data, types):
    pages=data.get("pages",[]) if isinstance(data,dict) else []
    out=[]
    for t in ("overview","placement","programme"):
        if t not in types: continue
        candidates=[p for p in pages if isinstance(p,dict) and p.get("type")==t]
        if candidates: out.append(candidates[0])
    return out


def text(v): return escape(str(v or "").strip())


def render_table(t):
    if not isinstance(t,dict) or not t.get("columns") or not t.get("rows"): return ""
    cols=t["columns"][:6]; rows=t["rows"][:20]
    s='<div class="table-wrapper"><table><thead><tr>'+''.join(f'<th>{text(c)}</th>' for c in cols)+'</tr></thead><tbody>'
    for row in rows:
        vals=row if isinstance(row,list) else [row.get(c,"") for c in cols]
        s+='<tr>'+''.join(f'<td>{text(v)}</td>' for v in vals[:len(cols)])+'</tr>'
    return s+'</tbody></table></div>'


def render_section(sec):
    s=f'<section class="main-section" id="{slug(sec.get("id") or sec.get("title"))}"><h2>{text(sec.get("title"))}</h2>'
    if sec.get("answer"): s+=f'<div class="answer-box"><strong>Quick answer</strong><p>{text(sec.get("answer"))}</p></div>'
    for p in sec.get("paragraphs",[])[:8]:
        if str(p).strip(): s+=f'<p>{text(p)}</p>'
    bullets=sec.get("bullets",[])[:12]
    if bullets: s+='<ul>'+''.join(f'<li>{text(x)}</li>' for x in bullets if str(x).strip())+'</ul>'
    s+=render_table(sec.get("table"))
    if sec.get("callout"):
        s+=f'<div class="pro-box"><h3>Important for applicants</h3><p>{text(sec.get("callout"))}</p></div>'
    return s+'</section>'


def render_page(page, college, rank, official_url, internal_links, filename):
    typ=page.get("type","")
    title=page.get("title") or f"{college} MBA"
    hero=page.get("hero_subtitle") or "MBA admission, eligibility, fees, selection and programme details."
    location=page.get("location_line") or official_url
    facts=page.get("facts") or [{"label":"Programme","value":"MBA / Management"},{"label":"Duration","value":"See official programme details"},{"label":"Mode","value":"See official programme details"},{"label":"Institute","value":college}]
    facts=(facts+[{}]*4)[:4]
    sections=page.get("sections") or []
    ids=[slug(x.get("id") or x.get("title")) for x in sections if isinstance(x,dict)]
    sidebar=''.join(f'<a href="#{i}">{text(next((x.get("title") for x in sections if slug(x.get("id") or x.get("title"))==i),i))}</a>' for i in ids)
    mobile=''.join(f'<a href="#{i}">{text(next((x.get("title") for x in sections if slug(x.get("id") or x.get("title"))==i),i))}</a>' for i in ids)
    quick=page.get("quick_answer") or (sections[0].get("answer") if sections else "")
    faqs=[x for x in page.get("faqs",[]) if isinstance(x,dict)][:8]
    sources=[]
    domain=urlparse(official_url).netloc.lower().replace("www.","")
    for u in page.get("source_urls",[]):
        if urlparse(str(u)).netloc.lower().replace("www.","")==domain and u not in sources: sources.append(u)
    sources=sources[:8]
    related=''.join(f'<a class="programme-link" href="{escape("/"+x)}">{text(Path(x).stem.replace("-"," ").title())}</a>' for x in internal_links[:4] if x!=filename)
    cta_url=page.get("cta_url") or official_url
    cta_label=page.get("cta_label") or "Visit official website"
    schema={"@context":"https://schema.org","@type":"WebPage","name":title,"description":hero,"url":f"https://vai2110.github.io/mba-admission-portal/{filename}","about":{"@type":"CollegeOrUniversity","name":college}}
    html='''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">'''
    html+=f'<title>{text(title)}</title><meta name="description" content="{text(hero)[:158]}"><link rel="canonical" href="https://vai2110.github.io/mba-admission-portal/{escape(filename)}"><meta property="og:title" content="{text(title)}"><meta property="og:description" content="{text(hero)[:158]}"><meta property="og:type" content="website"><link rel="stylesheet" href="college-page.css"><script type="application/ld+json">{json.dumps(schema,ensure_ascii=False)}</script></head><body>'
    html+='<header><div class="navbar"><a class="logo" href="index.html">MBA Admission Portal</a><nav class="nav-links"><a href="index.html">Home</a><a href="pgp.html">PGP</a></nav></div></header>'
    html+=f'<section class="hero"><div class="hero-container"><h1>{text(title)}</h1><h2>{text(hero)}</h2><div class="hero-location">{text(location)}</div></div></section>'
    html+='<div class="page-container"><div class="quick-facts">'+''.join(f'<div class="quick-fact"><span class="quick-fact-label">{text(x.get("label"))}</span><span class="quick-fact-value">{text(x.get("value"))}</span></div>' for x in facts)+'</div>'
    html+='<div class="mobile-on-page"><details><summary>On This Page</summary>'+mobile+'</details></div><div class="page-layout"><aside class="sidebar"><div class="sidebar-card"><div class="sidebar-title">On This Page</div>'+sidebar+'</div></aside><main class="content">'
    if quick: html+=f'<div class="answer-box"><strong>{"Quick answer"}</strong><p>{text(quick)}</p></div>'
    for sec in sections: html+=render_section(sec)
    if faqs:
        html+='<section class="main-section" id="faqs"><h2>FAQs</h2>'+''.join(f'<div class="faq-item"><h3>{text(x.get("question"))}</h3><p>{text(x.get("answer"))}</p></div>' for x in faqs)+'</section>'
    html+='<section class="main-section" id="next-steps"><h2>Next Steps</h2><div class="answer-box"><strong>Ready to check the latest official information?</strong><p>Use the official institute source before applying because dates, fees and admission criteria can change by cycle.</p></div><p><a class="cta" href="'+escape(str(cta_url))+'" target="_blank" rel="noopener">'+text(cta_label)+'</a></p>'
    if related: html+='<div class="programme-grid"><div class="programme-card"><span class="programme-label">Related pages</span><h3>Continue exploring</h3><p>'+related+'</p></div></div>'
    html+='</section><section class="main-section" id="official-sources"><h2>Official Sources</h2><p>Information on this page is based on the following official sources.</p><div class="official-links">'+''.join(f'<div class="official-link"><a href="{escape(str(u))}" target="_blank" rel="noopener">{text(str(u))}</a></div>' for u in sources)+'</div></section></main></div></div></body></html>'
    return html


def audit(html, page, official_url, existing, filename):
    soup=BeautifulSoup(html,"html.parser"); textv=soup.get_text(" ",strip=True); low=textv.lower(); issues=[]; score=0
    # Content quality 30
    if len(textv)>=6500: score+=12
    elif len(textv)>=4500: score+=10
    elif len(textv)>=3000: score+=7
    elif len(textv)>=2000: score+=4
    else: issues.append("content too thin")
    qhits=sum(k in low for k in {"admission","eligibility","fee","selection","cutoff","curriculum","placement","faq"})
    score+=min(8,qhits)
    if len(soup.find_all("table"))>=2: score+=5
    elif soup.find("table"): score+=3
    if len(soup.select(".faq-item"))>=5: score+=5
    # Architecture 25
    arch=[bool(soup.select_one("header .navbar .nav-links")),bool(soup.select_one(".hero .hero-container h1")),len(soup.select(".quick-fact"))==4,bool(soup.select_one(".page-layout .sidebar .sidebar-card")),bool(soup.select_one(".mobile-on-page")),len(soup.select(".main-section"))>=6,bool(soup.select_one(".answer-box")),bool(soup.select_one(".cta")),bool(soup.select_one(".official-links .official-link"))]
    score+=sum(2 for x in arch if x)
    if not all(arch): issues.append("reference architecture mismatch")
    # Technical / SEO 20
    tech=[len(soup.find_all("h1"))==1,bool(soup.title),bool(soup.find("meta",attrs={"name":"description"})),bool(soup.find("link",rel="canonical")),bool(soup.find("meta",attrs={"property":"og:title"})),bool(soup.find("meta",attrs={"property":"og:description"})),bool(soup.find("script",attrs={"type":"application/ld+json"})),bool(soup.find("meta",attrs={"name":"viewport"})),bool(soup.find("link",href=re.compile(r"college-page\.css")))]
    score+=sum(2 for x in tech if x)
    if not all(tech): issues.append("technical SEO integrity failure")
    # Source integrity 15
    domain=urlparse(official_url).netloc.lower().replace("www.","")
    src=[a.get("href") for a in soup.select(".official-link a[href]")]
    official=[u for u in src if urlparse(u).netloc.lower().replace("www.","")==domain]
    if len(official)>=4: score+=15
    elif len(official)>=2: score+=12
    elif official: score+=6
    else: issues.append("no official source links")
    # Hard failures
    if re.search(r"lorem ipsum|placeholder|insert here|xxx",textv,re.I): issues.append("placeholder content")
    if any(x in low for x in ("₹0","rs. 0","100% guaranteed")): issues.append("placeholder/unsupported claim")
    internal=[]
    for a in soup.find_all("a",href=True):
        h=a["href"]
        if h.startswith("/") and not h.startswith("//"): internal.append(h.lstrip("/"))
    broken=[x for x in internal if Path(ROOT/x).name not in existing and x!=filename]
    if broken: issues.append("broken internal links: "+", ".join(broken[:4]))
    critical=any(x in issues for x in ("placeholder content","placeholder/unsupported claim","no official source links","technical SEO integrity failure","reference architecture mismatch"))
    return min(score,100),issues,critical


def write_report(college, rank, score, pages, issues):
    p=ROOT/f"quality-audit-{slug(college)}-v8.json"
    p.write_text(json.dumps({"agent":"v8","college":college,"rank":rank,"score":score,"threshold":PUBLISH_THRESHOLD,"pages":pages,"issues":issues,"generated_at":time.strftime("%Y-%m-%dT%H:%M:%SZ",time.gmtime())},indent=2),encoding="utf-8")
    return p


def git(*args): subprocess.run(["git",*args],cwd=ROOT,check=True)


def main():
    if not SHEET_URL: raise RuntimeError("GOOGLE_SHEET_WEBAPP_URL is required")
    rows=read_master(); overrides=read_overrides(); selected=eligible_from_sheet()
    if not selected: print("No eligible colleges returned by Google Sheet."); return 0
    existing=existing_files()
    for meta in selected:
        rank=meta["rank"]; college=meta["college_name"]
        ov=overrides.get(rank,{})
        forced={x.strip() for x in str(ov.get("force_missing","")).split(",") if x.strip()}
        types=[x for x in ("overview","placement","programme") if x in (forced or set(missing_types(meta)))]
        if not types: continue
        print(f"=== V8 {rank}: {college} | {types} ===")
        sheet_post("updateStatus",rank=int(rank),researchStatus="Researching")
        url=discover_official(college,meta["official_website"])
        if not url:
            print("Blocked: official site unavailable"); continue
        research,source_pages=research(url)
        prompt=content_prompt(college,rank,url,research,types,existing)
        data=gemini(prompt); pages=normalize_content(data,types)
        if {p.get("type") for p in pages} != set(types):
            print("Blocked: model did not return every required page type"); continue
        final=[]; page_scores=[]; all_issues=[]
        for p in pages:
            fn=f"{slug(college)}-{slug(p.get('title') or p.get('type'))}.html"
            if p.get("type")=="overview": fn=f"{slug(college)}.html"
            elif p.get("type")=="placement": fn=f"{slug(college)}-placements.html"
            # programme filenames remain descriptive and college-specific
            if fn in existing or (ROOT/fn).exists():
                print(f"Protected collision: {fn}"); continue
            related=internal_targets(existing,{p.get("type")},college)
            html=render_page(p,college,rank,url,related,fn)
            sc,issues,critical=audit(html,p,url,existing,fn)
            # v8's content revision is a structured-content revision, not an HTML rewrite.
            for rev in range(MAX_REVISIONS):
                if sc>PUBLISH_THRESHOLD and not critical: break
                feedback=json.dumps(issues)
                revised=gemini(content_prompt(college,rank,url,research,[p.get("type")],existing)+f"\n\nREVISION REQUIRED. Fix these issues in the structured content: {feedback}\nReturn exactly one page of the requested type.")
                rp=normalize_content(revised,[p.get("type")])
                if not rp: break
                p=rp[0]; html=render_page(p,college,rank,url,related,fn); sc,issues,critical=audit(html,p,url,existing,fn)
                print(f"Revision {rev+1}: {fn} score={sc} critical={critical}")
            if sc<=PUBLISH_THRESHOLD or critical:
                all_issues += [f"{fn}: {x}" for x in issues]; print(f"NOT PUBLISHABLE: {fn} score={sc} issues={issues}"); continue
            final.append((fn,html,p)); page_scores.append(sc); all_issues += [f"{fn}: {x}" for x in issues]
        if not final or len(final)!=len(types):
            write_report(college,rank,round(sum(page_scores)/len(page_scores)) if page_scores else 0,[x[0] for x in final],all_issues); continue
        package=round(sum(page_scores)/len(page_scores)); report=write_report(college,rank,package,[x[0] for x in final],all_issues)
        passed=package>PUBLISH_THRESHOLD and not all_issues
        r=next((x for x in rows if str(x.get("rank"))==str(rank)),None)
        if r:
            r["quality_score"]=str(package); r["research_status"]="Research Complete"; r["qa_status"]="Passed" if passed else "Failed"
            for key,t in (("overview_status","overview"),("placement_status","placement"),("popular_course_status","programme")):
                if any(p.get("type")==t for _,_,p in final): r[key]="Done"
            r["deployment_status"]="Eligible - Auto Publish" if passed else "Blocked - QA <= 70 or critical failure"; r["live_verification"]="Not Started"
            write_master(rows); update_xlsx(rows)
        if not passed: continue
        for fn,html,_ in final: (ROOT/fn).write_text(html,encoding="utf-8")
        git("add",*[x[0] for x in final],str(report.relative_to(ROOT)),"data/college-content-master.csv","data/college-content-master.xlsx")
        git("commit","-m",f"Publish v8 benchmark-ready pages for {college}"); git("push","origin","main")
        if r: r["deployment_status"]="Pushed - Awaiting Pages"; write_master(rows); update_xlsx(rows)
        ok=[]
        for fn,_,_ in final:
            live=f"https://vai2110.github.io/mba-admission-portal/{fn}"; good=False
            for _ in range(8):
                try:
                    if session.get(live,timeout=20).status_code==200: good=True; break
                except requests.RequestException: pass
                time.sleep(10)
            ok.append(good)
        if r: r["live_verification"]="Verified" if all(ok) else "Failed"; write_master(rows); update_xlsx(rows)
    return 0

if __name__=="__main__": raise SystemExit(main())
