#!/usr/bin/env python3
"""Independent MBA college content production agent.

This agent is intentionally standalone. It does not read, import, execute or
follow AGENTS.md or any other repository agent configuration.
"""
import csv
import json
import os
import re
import subprocess
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "college-content-master.csv"
PRIORITY = ROOT / "data" / "college-content-priority.csv"
OVERRIDES = ROOT / "data" / "college-content-overrides.csv"
TRACKER = ROOT / "college-production-tracker.csv"
MODEL = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
API_KEY = os.getenv("GEMINI_API_KEY", "")
BATCH_SIZE = int(os.getenv("COLLEGE_BATCH_SIZE", "10"))
PUBLISH_THRESHOLD = 70
MAX_REVISION_PASSES = 2
TIMEOUT = 20

session = requests.Session()
session.headers.update({"User-Agent": "MBA-College-Content-Agent/2.0"})


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s or "").lower()).strip("-")


def fetch(url):
    try:
        r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
        if r.ok and "text/html" in r.headers.get("content-type", ""):
            return r.text, r.url
    except requests.RequestException:
        pass
    return "", url


def discover_official_url(college, supplied_url):
    if supplied_url:
        html, final = fetch(supplied_url)
        if html:
            return final
    try:
        r = requests.get(
            "https://www.google.com/search",
            params={"q": f'"{college}" official website'},
            timeout=TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        soup = BeautifulSoup(r.text, "html.parser")
        bad = ("wikipedia.org", "collegedunia.com", "shiksha.com", "careers360.com", "collegedekho.com", "getmyuni.com", "facebook.com", "linkedin.com")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("/url?q="):
                href = href.split("/url?q=", 1)[1].split("&", 1)[0]
            if href.startswith("http") and not any(x in href.lower() for x in bad):
                html, final = fetch(href)
                if html:
                    return final
    except requests.RequestException:
        pass
    return ""


def official_crawl(start_url, limit=28):
    domain = urlparse(start_url).netloc.lower().replace("www.", "")
    queue = [start_url]
    seen = set()
    pages = []
    keys = ("mba", "pgp", "pgdm", "management", "admission", "placement", "career", "fee", "fees", "programme", "program", "curriculum", "eligibility", "selection", "brochure", "prospectus", "annual", "report", "school", "department", "course")
    while queue and len(pages) < limit:
        url = queue.pop(0)
        if url in seen or urlparse(url).netloc.lower().replace("www.", "") != domain:
            continue
        seen.add(url)
        html, final = fetch(url)
        if not html:
            continue
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(" ", strip=True)
        if text:
            pages.append({"url": final, "title": soup.title.get_text(" ", strip=True) if soup.title else "", "text": text[:18000]})
        for a in soup.find_all("a", href=True):
            href = urljoin(final, a["href"]).split("#")[0]
            if urlparse(href).netloc.lower().replace("www.", "") != domain:
                continue
            hay = (a.get_text(" ", strip=True) + " " + href).lower()
            if any(k in hay for k in keys) and href not in seen and href not in queue:
                queue.append(href)
    return pages


def research_block(url):
    pages = official_crawl(url)
    return "\n\n--- OFFICIAL SOURCE ---\n\n".join(
        f"SOURCE URL: {p['url']}\nTITLE: {p['title']}\nCONTENT:\n{p['text']}" for p in pages
    ), pages


def gemini_json(prompt, max_tokens=30000):
    if not API_KEY:
        raise RuntimeError("GEMINI_API_KEY is missing")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent"
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.15, "maxOutputTokens": max_tokens, "responseMimeType": "application/json"},
    }
    last = None
    for attempt in range(4):
        try:
            r = session.post(url, params={"key": API_KEY}, json=payload, timeout=180)
            if r.status_code in (429, 500, 502, 503, 504):
                last = r
                time.sleep(5 * (attempt + 1))
                continue
            r.raise_for_status()
            return json.loads(r.json()["candidates"][0]["content"]["parts"][0]["text"])
        except requests.RequestException as e:
            last = e
            if attempt < 3:
                time.sleep(5 * (attempt + 1))
            else:
                raise
    raise RuntimeError(f"Gemini unavailable after retries: {last}")


def response_pages(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("pages", [])
    return []


def read_master():
    with MASTER.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def priority_rows(rows):
    if not PRIORITY.exists():
        return rows
    by_rank = {str(r.get("rank")): r for r in rows}
    out = []
    with PRIORITY.open(encoding="utf-8-sig", newline="") as f:
        for p in csv.DictReader(f):
            r = by_rank.get(str(p.get("rank")))
            if r:
                r["college_name"] = p["college_name"]
                r["official_website"] = p["official_website"]
                out.append(r)
    return out + [r for r in rows if r not in out]


def read_overrides():
    if not OVERRIDES.exists():
        return {}
    with OVERRIDES.open(encoding="utf-8-sig", newline="") as f:
        return {str(r.get("rank")): r for r in csv.DictReader(f)}


def apply_overrides(rows):
    ov = read_overrides()
    for r in rows:
        x = ov.get(str(r.get("rank")))
        if x:
            for k in ("overview_status", "placement_status", "popular_course_status", "quality_score", "research_status", "qa_status", "deployment_status", "live_verification"):
                if x.get(k) != "":
                    r[k] = x[k]
    return ov


def read_tracker():
    if not TRACKER.exists():
        return []
    with TRACKER.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def tracker_for(college, tracker):
    def keys(s):
        raw = str(s or "").lower()
        v = {norm(raw)}
        for a, b in [("indian institute of management", "iim"), ("indian institute of technology", "iit"), ("national institute of technology", "nit"), ("institute of management technology", "imt"), ("management development institute", "mdi"), ("symbiosis institute of business management", "sibm"), ("t. a. pai management institute", "tapmi"), ("great lakes institute of management", "great lakes"), ("chandigarh university", "cu")]:
            if a in raw:
                v.add(norm(raw.replace(a, b)))
        return v
    tk = keys(college)
    for r in tracker:
        rk = keys(r.get("College Name", ""))
        if tk & rk:
            return r
        if any(len(a) >= 4 and any(a in b or b in a for b in rk) for a in tk):
            return r
    return None


def missing_types(college, tracker, forced=None):
    if forced:
        return forced
    r = tracker_for(college, tracker)
    if not r:
        return {"overview", "placement", "programme"}
    def exists(v):
        return str(v or "").strip().lower() not in ("", "pending", "not started", "not started - pending", "n/a", "not applicable")
    m = set()
    if not exists(r.get("Overview Page")):
        m.add("overview")
    if not exists(r.get("Placement Page")):
        m.add("placement")
    courses = [r.get(k, "") for k in ("Course Page 1", "Course Page 2", "Course Page 3")]
    if courses and not any(exists(v) for v in courses) and not all(str(v or "").strip().lower() == "not applicable" for v in courses):
        m.add("programme")
    return m


def generation_prompt(college, rank, url, research, types, feedback=""):
    return f"""You are an independent MBA admissions content generator. College: {college}. NIRF 2025 rank: {rank}. Official domain: {url}. Generate ONLY these page types: {', '.join(sorted(types))}.
Use ONLY supplied official-source material for factual claims. Never invent or infer fees, dates, seats, salaries, cutoffs, eligibility, rankings, recruiters, programme names or statistics. If unavailable, explicitly state that the official source does not publish it. Reddit/Quora are permitted ONLY for cutoff information when official cutoff data is absent and evidence is supplied.
Use the approved IIM Ahmedabad and SIBM Pune reference architecture, hierarchy, content depth and responsive presentation as design/quality references only. Use shared college-page.css. Do not copy benchmark wording or facts.
Student-centric, simple English, answer-first. No generic AI intro, keyword stuffing, filler, repetition or unsupported marketing claims. Tables where useful. Include clear admission, eligibility, fees/fee availability, selection, curriculum/programme information, placements where relevant, FAQs, CTAs and internal links where supported.
Every page MUST have a college-specific filename. Never use generic filenames such as programme.html, overview.html, placement.html, mba.html or index.html. Filename pattern: {slugify(college)}-{{descriptive-page-slug}}.html.
HTML must be a complete HTML5 document with viewport, one H1, logical H2/H3, title, meta description, canonical, OG title/description, JSON-LD, FAQ content, official source links and shared college-page.css. Internal links may target only known existing/package files. Do not invent URLs. Return JSON ONLY: {{\"pages\":[{{\"filename\":\"{slugify(college)}-mba-programme.html\",\"type\":\"overview|placement|programme\",\"title\":\"...\",\"html\":\"<complete HTML5 document>\",\"source_urls\":[\"official URL\"]}}]}}.
Revision feedback to fix, if any: {feedback}

OFFICIAL SOURCE MATERIAL:\n{research[:180000]}"""


def revision_prompt(college, rank, url, research, types, failures, previous):
    return f"""Perform a targeted quality revision for {college}, NIRF rank {rank}. Generate ONLY page types: {', '.join(sorted(types))}.
Use ONLY supplied official sources. Fix every listed audit issue. Do not merely rewrite prose. Preserve verified facts and remove unsupported claims.
MANDATORY checklist: complete HTML5; viewport; college-specific filename; one H1; 3+ useful H2 sections; answer-first opening; useful tables; student-intent sections; title and meta description; canonical; OG metadata; valid JSON-LD matching the page; college-specific FAQs with grounded answers; official source section with 2+ official URLs where available; internal links only to existing/package files; shared college-page.css; clear CTA; no placeholders, lorem ipsum, fake numbers or generic claims.
Quality target: score above 70 and zero critical failures.
Audit issues: {json.dumps(failures)}
Previous pages: {json.dumps(previous)[:130000]}
OFFICIAL SOURCES:\n{research[:160000]}"""


def _page_filename(college, page, used):
    raw = Path(str(page.get("filename", "")).strip()).name
    stem = slugify(Path(raw).stem)
    college_slug = slugify(college)
    generic = {"programme", "program", "overview", "placement", "mba", "index", "page"}
    if not stem or stem in generic or not stem.startswith(college_slug):
        title_slug = slugify(page.get("title", ""))
        if page.get("type") == "programme":
            desc = title_slug or "mba-programme"
            if "mba" not in desc and "pgdm" not in desc and "pgp" not in desc:
                desc = "mba-" + desc
        elif page.get("type") == "overview":
            desc = "overview"
        else:
            desc = "placements"
        stem = f"{college_slug}-{desc}"
    candidate = stem + ".html"
    i = 2
    while candidate in used:
        candidate = f"{stem}-{i}.html"
        i += 1
    return candidate


def normalize_pages(pages, college):
    used = set()
    mapping = {}
    normalized = []
    for page in pages:
        if not isinstance(page, dict) or not page.get("html") or page.get("type") not in {"overview", "placement", "programme"}:
            continue
        old = Path(str(page.get("filename", "")).strip()).name
        new = _page_filename(college, page, used)
        used.add(new)
        mapping[old] = new
        p = dict(page)
        p["filename"] = new
        normalized.append(p)
    if mapping:
        for p in normalized:
            html = p["html"]
            for old, new in mapping.items():
                if old:
                    html = re.sub(rf'(["\'(]){re.escape(old)}(["\')#?])', rf'\1{new}\2', html)
            p["html"] = html
    return normalized


def _internal_links(soup):
    links = []
    for a in soup.find_all("a", href=True):
        href = str(a.get("href", "")).strip()
        if href and not href.startswith(("http://", "https://", "#", "mailto:", "tel:", "javascript:")):
            links.append(href.split("#")[0].split("?")[0].lstrip("/"))
    return [x for x in links if x]


def audit(html, source_urls, official_url, files, page_type=""):
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    low = text.lower()
    critical = []
    notes = []
    score = 0

    # 20: content usefulness
    if len(text) >= 2500: score += 8
    elif len(text) >= 1800: score += 6
    elif len(text) >= 1000: score += 3
    else: notes.append("low content depth")
    intent = {
        "overview": ("admission", "programme", "placement"),
        "placement": ("placement", "career", "recruit"),
        "programme": ("eligibility", "admission", "fee", "selection", "curriculum"),
    }.get(page_type, ("admission", "programme", "placement"))
    hits = sum(1 for k in intent if k in low)
    score += min(7, hits * 2)
    if re.search(r"what|how|when|who|which|why", low[:1200]): score += 2
    if len(re.findall(r"\b(the|and|for|with)\b", low)) > 30: score += 3

    # 20: source integrity
    domain = urlparse(official_url).netloc.lower().replace("www.", "")
    official_sources = [u for u in source_urls if urlparse(str(u)).netloc.lower().replace("www.", "") == domain]
    if len(official_sources) >= 3: score += 10
    elif len(official_sources) >= 2: score += 8
    elif official_sources: score += 4
    else: critical.append("no official sources")
    if len(official_sources) != len(source_urls): critical.append("non-official source URL")
    if soup.find(string=re.compile("official source", re.I)): score += 5
    if soup.find_all("a", href=re.compile(r"^https?://")) and official_sources: score += 5

    # 20: technical SEO
    if soup.find("h1"): score += 4
    else: critical.append("missing H1")
    if len(soup.find_all("h2")) >= 3: score += 4
    elif soup.find_all("h2"): score += 2
    else: notes.append("thin heading structure")
    if soup.title and soup.title.get_text(strip=True): score += 2
    else: critical.append("missing title")
    if soup.find("meta", attrs={"name": "description"}): score += 2
    else: critical.append("missing meta description")
    if soup.find("link", rel="canonical"): score += 3
    else: critical.append("missing canonical")
    if soup.find("meta", attrs={"property": "og:title"}) and soup.find("meta", attrs={"property": "og:description"}): score += 2
    else: notes.append("incomplete OG metadata")
    if soup.find("script", attrs={"type": "application/ld+json"}): score += 3
    else: notes.append("missing JSON-LD")

    # 20: UX and internal links
    if soup.find("table"): score += 4
    if len(soup.find_all(class_=re.compile(r"faq", re.I))) >= 3 or low.count("faq") >= 3: score += 4
    internal = _internal_links(soup)
    broken = [x for x in internal if Path(ROOT / x).name not in files and not x.startswith("#")]
    if not broken: score += 6
    else: critical.append("broken internal link(s): " + ", ".join(broken[:5]))
    if "college-page.css" in str(html): score += 4
    else: notes.append("shared college-page.css missing")
    if soup.find("meta", attrs={"name": "viewport"}): score += 2

    # 20: completeness/student experience
    required = {
        "overview": ("admission", "programme", "placement", "fee"),
        "placement": ("placement", "recruit", "career"),
        "programme": ("eligibility", "admission", "fee", "selection", "curriculum"),
    }.get(page_type, ("admission", "programme"))
    score += min(10, sum(2 for k in required if k in low))
    if soup.find("a", href=re.compile(r"apply|admission|contact", re.I)): score += 3
    if soup.find_all("a", href=True): score += 2
    if "student" in low or "applicant" in low: score += 2
    if "lorem" in low or "placeholder" in low: critical.append("placeholder content")
    if any(x in low for x in ("₹0", "rs. 0", "to be updated", "insert here")): critical.append("placeholder/zero value")
    if re.search(r"\b(guaranteed|best in india|100% placement)\b", low): notes.append("potentially unsupported marketing language")

    return min(score, 100), critical, notes


def write_master(rows):
    with MASTER.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def update_xlsx(rows):
    path = ROOT / "data" / "college-content-master.xlsx"
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Management"
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, 1): ws.cell(1, c, h)
    for rr, r in enumerate(rows, 2):
        for c, h in enumerate(headers, 1): ws.cell(rr, c, r.get(h, ""))
    ws.freeze_panes = "A2"
    wb.save(path)


def git(*args):
    subprocess.run(["git", *args], cwd=ROOT, check=True)


def main():
    rows = read_master()
    overrides = apply_overrides(rows)
    tracker = read_tracker()
    existing = {p.name for p in ROOT.glob("*.html")}
    generated = 0

    for r in priority_rows(rows):
        if generated >= BATCH_SIZE:
            break
        college = r["college_name"].strip()
        ov = overrides.get(str(r.get("rank")), {})
        forced = {x.strip() for x in str(ov.get("force_missing", "") or "").split(",") if x.strip()}
        types = missing_types(college, tracker, forced)
        if not types:
            r["qa_status"] = "Skipped - existing pages protected"
            r["deployment_status"] = "Already Exists"
            r["live_verification"] = "Checked"
            write_master(rows); update_xlsx(rows)
            continue

        r["research_status"] = "Researching"
        write_master(rows); update_xlsx(rows)
        print(f"=== {r['rank']}: {college} | {sorted(types)} ===")
        url = discover_official_url(college, r["official_website"])
        if not url:
            r["research_status"] = "Blocked - official site not discovered"
            r["qa_status"] = "Failed"
            r["quality_score"] = "0"
            write_master(rows); update_xlsx(rows)
            generated += 1
            continue

        r["official_website"] = url
        research, source_pages = research_block(url)
        r["research_status"] = "Research Complete"
        write_master(rows); update_xlsx(rows)

        data = gemini_json(generation_prompt(college, r["rank"], url, research, types))
        pages = normalize_pages(response_pages(data), college)
        safe = []
        for p in pages:
            if p.get("type") not in types:
                continue
            fn = Path(p.get("filename", "")).name
            if fn and fn not in existing and not (ROOT / fn).exists():
                safe.append((fn, p))
        if not safe:
            r["qa_status"] = "Failed - generator returned no missing page"
            r["quality_score"] = "0"
            write_master(rows); update_xlsx(rows)
            generated += 1
            continue

        scores, critical_all, notes = [], [], []
        for fn, p in safe:
            sc, cr, no = audit(p["html"], p.get("source_urls", []), url, existing | {x[0] for x in safe}, p.get("type", ""))
            scores.append(sc); critical_all += [f"{fn}: {x}" for x in cr]; notes += [f"{fn}: {x}" for x in no]
            (ROOT / fn).write_text(p["html"], encoding="utf-8")
        package_score = round(sum(scores) / len(scores))

        # Up to two targeted revisions. A score <=70 OR any critical issue is never publishable.
        for revision_no in range(1, MAX_REVISION_PASSES + 1):
            if package_score > PUBLISH_THRESHOLD and not critical_all:
                break
            feedback = critical_all + notes + [f"Current package score: {package_score}; target > {PUBLISH_THRESHOLD}."]
            revision = gemini_json(revision_prompt(college, r["rank"], url, research, types, feedback, pages))
            revised = normalize_pages([p for p in response_pages(revision) if p.get("type") in types and p.get("html")], college)
            if not revised:
                notes.append(f"revision {revision_no} returned no usable pages")
                break
            candidate = []
            used = set(existing)
            for p in revised:
                fn = Path(p.get("filename", "")).name
                if fn and fn not in used:
                    candidate.append((fn, p)); used.add(fn)
            if not candidate:
                notes.append(f"revision {revision_no} returned only protected/colliding filenames")
                break
            safe = candidate
            pages = [p for _, p in safe]
            scores, critical_all, notes = [], [], []
            package = {x[0] for x in safe}
            for fn, p in safe:
                sc, cr, no = audit(p["html"], p.get("source_urls", []), url, existing | package, p.get("type", ""))
                scores.append(sc); critical_all += [f"{fn}: {x}" for x in cr]; notes += [f"{fn}: {x}" for x in no]
                (ROOT / fn).write_text(p["html"], encoding="utf-8")
            package_score = round(sum(scores) / len(scores))
            print(f"Revision {revision_no}: {college} score={package_score}")

        r["quality_score"] = str(package_score)
        passed = package_score > PUBLISH_THRESHOLD and not critical_all
        r["qa_status"] = "Passed" if passed else "Failed"
        for k, t in (("overview_status", "overview"), ("placement_status", "placement"), ("popular_course_status", "programme")):
            if any(p.get("type") == t for _, p in safe):
                r[k] = "Done"
        report = {
            "college": college,
            "rank": r["rank"],
            "score": package_score,
            "threshold": PUBLISH_THRESHOLD,
            "page_scores": scores,
            "critical_failures": critical_all,
            "notes": notes,
            "source_count": len(source_pages),
            "generated_pages": [x[0] for x in safe],
            "revision_passes": MAX_REVISION_PASSES,
        }
        report_path = ROOT / f"quality-audit-{slugify(college)}.json"
        report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
        r["deployment_status"] = "Not Started"
        r["live_verification"] = "Not Started"
        write_master(rows); update_xlsx(rows)

        git("config", "user.email", "41898282+github-actions[bot]@users.noreply.github.com")
        git("config", "user.name", "github-actions[bot]")
        if passed:
            r["deployment_status"] = "Eligible - Auto Publish"
            write_master(rows); update_xlsx(rows)
            git("add", *[fn for fn, _ in safe], "data/college-content-master.csv", "data/college-content-master.xlsx", str(report_path.relative_to(ROOT)))
            git("commit", "-m", f"Publish independently generated pages for {college}")
            git("push", "origin", "main")
            r["deployment_status"] = "Pushed - Awaiting Pages"
            write_master(rows); update_xlsx(rows)
            urls = []
            for fn, _ in safe:
                live = f"https://vai2110.github.io/mba-admission-portal/{fn}"
                ok = False
                for _ in range(8):
                    try:
                        if session.get(live, timeout=20).status_code == 200:
                            ok = True
                            break
                    except requests.RequestException:
                        pass
                    time.sleep(10)
                urls.append(ok)
            r["live_verification"] = "Verified" if all(urls) else "Failed"
            write_master(rows); update_xlsx(rows)
            git("add", "data/college-content-master.csv", "data/college-content-master.xlsx")
            git("commit", "-m", f"Update quality and live status for {college}")
            git("push", "origin", "main")
        else:
            r["deployment_status"] = "Blocked - QA <= 70 or critical failure"
            for fn, _ in safe:
                try:
                    (ROOT / fn).unlink()
                except FileNotFoundError:
                    pass
            write_master(rows); update_xlsx(rows)
            git("add", "data/college-content-master.csv", "data/college-content-master.xlsx", str(report_path.relative_to(ROOT)))
            git("commit", "-m", f"Hold {college} pages for QA revision")
            git("push", "origin", "main")
        existing.update(x[0] for x in safe)
        generated += 1


if __name__ == "__main__":
    main()
